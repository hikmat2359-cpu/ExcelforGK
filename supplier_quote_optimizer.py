import streamlit as st
import pandas as pd
from io import BytesIO
import zipfile
import os
from datetime import datetime
from auth_utils import (
    initialize_session_state, 
    require_authentication, 
    show_login_form, 
    show_user_info,
    show_user_management,
    is_session_valid
)

st.set_page_config(page_title="Supplier Quote Optimizer", layout="wide")

# Initialize authentication
initialize_session_state()

# Check authentication before showing the app
if not is_session_valid():
    show_login_form()
    st.stop()

# Show user info and logout option
show_user_info()
st.markdown("---")

# Initialize session state for user selections and processing state
if 'supplier_selections' not in st.session_state:
    st.session_state.supplier_selections = {}
if 'processing_complete' not in st.session_state:
    st.session_state.processing_complete = False
if 'optimized_allocation' not in st.session_state:
    st.session_state.optimized_allocation = []

st.title("📦 Supplier Quote Optimizer")
st.write(
    "Upload your **Order List** and **Supplier Quotes** Excel files. "
    "The app will automatically optimize supplier selection based on price and availability. "
    "You can then manually override selections and download optimized files grouped by supplier."
)

# -----------------------------
# Excel Sheet Creation Functions
# -----------------------------

def create_price_comparison_sheet(order_df, quote_df, final_df):
    """Create Sheet 1: Order list with price comparison columns showing separate rows for each allocation"""
    
    # Create a list to store all allocation rows
    allocation_rows = []
    
    # Get all suppliers for price comparison columns
    all_suppliers = sorted(quote_df['Supplier'].unique())
    
    # Process each part in the order
    for _, order_row in order_df.iterrows():
        part_number = order_row['PartNumber']
        qty_required = order_row['QtyRequired']
        
        # Get all allocations for this part from final_df
        part_allocations = final_df[final_df['PartNumber'] == part_number]
        
        if part_allocations.empty:
            # No allocation found - create a single row showing not allocated
            row_data = {
                'Part Number': part_number,
                'Qty Required': qty_required,
                'Allocated Qty': 0,
                'Selected Supplier': 'Not Selected',
                'Selected Price': 'N/A',
                'Total Cost': 0,
                'Allocation Source': 'Not Allocated'
            }
            
            # Add price columns for all suppliers
            for supplier in all_suppliers:
                supplier_quotes = quote_df[(quote_df['Supplier'] == supplier) & (quote_df['PartNumber'] == part_number)]
                # Limit supplier name to 10 characters for column headers
                supplier_short = str(supplier)[:10]
                price_col = f"{supplier_short}_Price"
                qty_col = f"{supplier_short}_Qty"
                
                if not supplier_quotes.empty:
                    supplier_quote = supplier_quotes.iloc[0]
                    row_data[price_col] = f"{supplier_quote['UnitPrice']:.2f}"
                    row_data[qty_col] = supplier_quote['AvailableQty']
                else:
                    row_data[price_col] = "N/A"
                    row_data[qty_col] = "N/A"
            
            allocation_rows.append(row_data)
        else:
            # Create separate rows for each allocation of this part
            for _, allocation in part_allocations.iterrows():
                row_data = {
                    'Part Number': part_number,
                    'Qty Required': qty_required,
                    'Allocated Qty': allocation['QtyAllocated'],
                    'Selected Supplier': allocation['Supplier'],
                    'Selected Price': f"{allocation['UnitPrice']:.2f}" if allocation['UnitPrice'] > 0 else "N/A",
                    'Total Cost': allocation['TotalCost'],
                    'Allocation Source': allocation['AllocationSource']
                }
                
                # Add price columns for all suppliers
                for supplier in all_suppliers:
                    supplier_quotes = quote_df[(quote_df['Supplier'] == supplier) & (quote_df['PartNumber'] == part_number)]
                    # Limit supplier name to 10 characters for column headers
                    supplier_short = str(supplier)[:10]
                    price_col = f"{supplier_short}_Price"
                    qty_col = f"{supplier_short}_Qty"
                    
                    if not supplier_quotes.empty:
                        supplier_quote = supplier_quotes.iloc[0]
                        row_data[price_col] = f"{supplier_quote['UnitPrice']:.2f}"
                        row_data[qty_col] = supplier_quote['AvailableQty']
                    else:
                        row_data[price_col] = "N/A"
                        row_data[qty_col] = "N/A"
                
                allocation_rows.append(row_data)
    
    # Create DataFrame from allocation rows
    result_df = pd.DataFrame(allocation_rows)
    
    # Ensure columns are in the right order
    base_columns = ['Part Number', 'Qty Required', 'Allocated Qty', 'Selected Supplier', 'Selected Price', 'Total Cost', 'Allocation Source']
    supplier_columns = []
    for supplier in all_suppliers:
        # Limit supplier name to 10 characters for column headers
        supplier_short = str(supplier)[:10]
        supplier_columns.extend([f"{supplier_short}_Price", f"{supplier_short}_Qty"])
    
    # Reorder columns
    all_columns = base_columns + supplier_columns
    result_df = result_df.reindex(columns=all_columns, fill_value="N/A")
    
    # Add sum totals at the bottom for Allocated Qty and Total Cost only
    if not result_df.empty:
        # Calculate sums for numeric columns (excluding Qty Required)
        allocated_qty_sum = pd.to_numeric(result_df['Allocated Qty'], errors='coerce').sum()
        total_cost_sum = pd.to_numeric(result_df['Total Cost'], errors='coerce').sum()
        
        # Create totals row
        totals_row = pd.Series(index=result_df.columns, dtype=object)
        totals_row['Part Number'] = 'TOTALS'
        totals_row['Qty Required'] = ''  # Empty for Qty Required
        totals_row['Allocated Qty'] = allocated_qty_sum
        totals_row['Selected Supplier'] = ''
        totals_row['Selected Price'] = ''
        totals_row['Total Cost'] = total_cost_sum
        totals_row['Allocation Source'] = ''
        
        # Fill supplier columns with empty strings
        for supplier in all_suppliers:
            # Limit supplier name to 10 characters for column headers
            supplier_short = str(supplier)[:10]
            totals_row[f"{supplier_short}_Price"] = ''
            totals_row[f"{supplier_short}_Qty"] = ''
        
        # Add empty row before totals
        empty_row = pd.Series(index=result_df.columns, dtype=object)
        empty_row = empty_row.fillna('')
        
        # Append empty row and totals row to the DataFrame
        result_df = pd.concat([result_df, pd.DataFrame([empty_row]), pd.DataFrame([totals_row])], ignore_index=True)
    
    return result_df

def apply_excel_highlighting(writer, sheet_name, sheet_data, final_df):
    """Apply highlighting to selected prices in Excel using openpyxl with different colors for manual vs auto selections"""
    from openpyxl.styles import PatternFill, Font, Border, Side
    
    workbook = writer.book
    worksheet = workbook[sheet_name]
    
    # Define highlighting styles for different allocation sources
    manual_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold for manual selections
    auto_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green for auto-optimized
    partial_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange for partial manual
    
    manual_font = Font(bold=True, color="8B4513")  # Dark brown text for manual
    auto_font = Font(bold=True, color="006400")    # Dark green text for auto
    partial_font = Font(bold=True, color="FF4500")  # Red-orange text for partial
    
    manual_border = Border(
        left=Side(border_style="thick", color="DAA520"),
        right=Side(border_style="thick", color="DAA520"),
        top=Side(border_style="thick", color="DAA520"),
        bottom=Side(border_style="thick", color="DAA520")
    )
    
    auto_border = Border(
        left=Side(border_style="thin", color="008000"),
        right=Side(border_style="thin", color="008000"),
        top=Side(border_style="thin", color="008000"),
        bottom=Side(border_style="thin", color="008000")
    )
    
    partial_border = Border(
        left=Side(border_style="medium", color="FF4500"),
        right=Side(border_style="medium", color="FF4500"),
        top=Side(border_style="medium", color="FF4500"),
        bottom=Side(border_style="medium", color="FF4500")
    )
    
    # Get column indices for supplier price columns
    supplier_price_cols = {}
    for col_idx, col_name in enumerate(sheet_data.columns, 1):
        if col_name.endswith('_Price'):
            supplier_name = col_name.replace('_Price', '')
            supplier_price_cols[supplier_name] = col_idx
    
    # Apply highlighting to selected prices based on allocation source
    for row_idx, row in sheet_data.iterrows():
        part_number = row['Part Number']
        selected_supplier = row['Selected Supplier']
        
        # Check both full supplier name and shortened version for highlighting
        supplier_match = None
        if selected_supplier != "Not Selected":
            # First try exact match
            if selected_supplier in supplier_price_cols:
                supplier_match = selected_supplier
            else:
                # Try shortened version (first 10 characters)
                supplier_short = str(selected_supplier)[:10]
                if supplier_short in supplier_price_cols:
                    supplier_match = supplier_short
        
        if supplier_match:
            col_idx = supplier_price_cols[supplier_match]
            cell = worksheet.cell(row=row_idx + 2, column=col_idx)  # +2 for header and 1-based indexing
            
            # Determine allocation source for this part and supplier
            part_allocations = final_df[
                (final_df['PartNumber'] == part_number) & 
                (final_df['Supplier'] == selected_supplier)
            ]
            
            if not part_allocations.empty:
                allocation_source = part_allocations.iloc[0]['AllocationSource']
                
                # Apply different styling based on allocation source
                if 'Manual Selection' in allocation_source:
                    if 'Partial' in allocation_source:
                        cell.fill = partial_fill
                        cell.font = partial_font
                        cell.border = partial_border
                    else:
                        cell.fill = manual_fill
                        cell.font = manual_font
                        cell.border = manual_border
                else:  # Auto-Optimized
                    cell.fill = auto_fill
                    cell.font = auto_font
                    cell.border = auto_border
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Add padding and cap at 50
        worksheet.column_dimensions[column_letter].width = adjusted_width

def create_supplier_groups_sheet(writer, final_df, quote_df):
    """Create Sheet 2: Rows grouped by suppliers"""
    
    if final_df.empty:
        # Create an empty sheet if no data
        empty_df = pd.DataFrame({'Message': ['No supplier data available']})
        empty_df.to_excel(writer, sheet_name='Supplier_Groups', index=False)
        return
    
    # Convert all suppliers to strings before sorting to handle mixed data types
    suppliers = sorted([str(s) for s in final_df['Supplier'].unique()])
    
    if not suppliers:
        # Create an empty sheet if no suppliers
        empty_df = pd.DataFrame({'Message': ['No suppliers found']})
        empty_df.to_excel(writer, sheet_name='Supplier_Groups', index=False)
        return
    
    for supplier in suppliers:
        # Convert supplier to string for consistent comparison
        supplier_str = str(supplier)
        supplier_data = final_df[final_df['Supplier'].astype(str) == supplier_str].copy()
        
        if supplier_data.empty:
            continue
            
        # Add additional supplier info
        supplier_quotes = quote_df[quote_df['Supplier'].astype(str) == supplier_str]
        supplier_data = supplier_data.merge(
            supplier_quotes[['PartNumber', 'AvailableQty']],
            on='PartNumber',
            how='left'
        )
        
        # Rename columns for clarity
        supplier_data = supplier_data.rename(columns={
            'PartNumber': 'Part Number',
            'QtyAllocated': 'Qty Ordered',
            'UnitPrice': 'Unit Price',
            'TotalCost': 'Total Cost',
            'AvailableQty': 'Available Qty'
        })
        
        # Select relevant columns
        supplier_data = supplier_data[['Part Number', 'Qty Ordered', 'Unit Price', 'Available Qty', 'Total Cost']]
        
        # Create sheet name (Excel sheet names have 31 char limit and cannot contain certain characters)
        # Remove invalid characters: / \ ? * [ ] : ( )
        clean_supplier = supplier_str.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_').replace('(', '_').replace(')', '_')
        sheet_name = f"Supplier_{clean_supplier[:20]}"
        supplier_data.to_excel(writer, sheet_name=sheet_name, index=False)

def create_not_available_sheet(order_df, quote_df):
    """Create Sheet 3: Parts that are not available from any supplier"""
    
    # Find parts in order that are not in quotes
    quoted_parts = set(quote_df['PartNumber'].unique())
    ordered_parts = set(order_df['PartNumber'].unique())
    
    not_available_parts = ordered_parts - quoted_parts
    
    if not_available_parts:
        not_available_df = order_df[order_df['PartNumber'].isin(not_available_parts)].copy()
        not_available_df = not_available_df.rename(columns={
            'PartNumber': 'Part Number',
            'QtyRequired': 'Qty Required'
        })
        not_available_df['Status'] = 'No quotes available'
        return not_available_df
    else:
        return pd.DataFrame()

def create_combined_suppliers_sheet(writer, final_df, quote_df, order_df):
    """Create Sheet 2: Combined sheet with supplier-wise summaries, subtotals, grand totals, and not available items"""
    from openpyxl.styles import PatternFill, Font, Border, Side
    
    combined_data = []
    grand_total_qty = 0
    grand_total_amount = 0
    na_header_row = None
    subtotal_na_row = None
    total_qty_row = None
    grand_total_row = None
    subtotal_rows = []  # Track all subtotal rows for formatting
    supplier_header_rows = []  # Track supplier header rows for formatting
    column_header_rows = []  # Track column header rows for formatting
    
    if not final_df.empty:
        # Convert all suppliers to strings before sorting to handle mixed data types
        suppliers = sorted([str(s) for s in final_df['Supplier'].unique()])
        
        for supplier in suppliers:
            # Skip special status entries including shortage items
            if supplier in ['NOT AVAILABLE', 'SHORTAGE', 'N/A (SHORTAGE)']:
                continue
                
            # Convert supplier to string for consistent comparison
            supplier_str = str(supplier)
            supplier_data = final_df[final_df['Supplier'].astype(str) == supplier_str].copy()
            
            if supplier_data.empty:
                continue
            
            # Create dynamic supplier name limited to 10 digits
            supplier_name = supplier_str[:10]  # Limit to 10 characters
            
            # Add supplier header row with dynamic name
            supplier_header_rows.append(len(combined_data))  # Track row index for formatting
            combined_data.append({
                'Part Number': supplier_name,
                'Qty Ordered': '',
                'Unit Price': '',
                'Total Cost': ''
            })
            
            # Add column header row above each supplier table
            column_header_rows.append(len(combined_data))  # Track row index for formatting
            combined_data.append({
                'Part Number': 'Part Number',
                'Qty Ordered': 'Qty Ordered',
                'Unit Price': 'Unit Price',
                'Total Cost': 'Total Cost'
            })
            
            # Process supplier data and calculate subtotals
            supplier_total_qty = 0
            supplier_total_amount = 0
            
            for _, row in supplier_data.iterrows():
                qty = row['QtyAllocated']
                total_cost = row['TotalCost']
                
                combined_data.append({
                    'Part Number': row['PartNumber'],
                    'Qty Ordered': qty,
                    'Unit Price': row['UnitPrice'],
                    'Total Cost': total_cost
                })
                
                supplier_total_qty += qty
                supplier_total_amount += total_cost
            
            # Add supplier subtotal row
            subtotal_rows.append(len(combined_data))  # Track for formatting
            combined_data.append({
                'Part Number': 'SUBTOTAL',
                'Qty Ordered': supplier_total_qty,
                'Unit Price': '',
                'Total Cost': supplier_total_amount
            })
            
            # Add empty row after each supplier
            combined_data.append({
                'Part Number': '',
                'Qty Ordered': '',
                'Unit Price': '',
                'Total Cost': ''
            })
            
            # Add to grand totals
            grand_total_qty += supplier_total_qty
            grand_total_amount += supplier_total_amount
    
    # Add grand totals section if there are suppliers
    if grand_total_qty > 0:
        grand_total_row = len(combined_data)  # Track for formatting
        combined_data.append({
            'Part Number': 'GRAND TOTAL',
            'Qty Ordered': grand_total_qty,
            'Unit Price': '',
            'Total Cost': grand_total_amount
        })
        
        # Add empty row before not available section
        combined_data.append({
            'Part Number': '',
            'Qty Ordered': '',
            'Unit Price': '',
            'Total Cost': ''
        })
    
    # Add consolidated N/A section (includes both not available parts and shortage items)
    not_available_df = create_not_available_sheet(order_df, quote_df)
    not_available_total_qty = 0
    
    # Get shortage items from final_df (items with 'N/A (SHORTAGE)' supplier)
    shortage_items = []
    if not final_df.empty:
        shortage_data = final_df[final_df['Supplier'] == 'N/A (SHORTAGE)']
        if not shortage_data.empty:
            for _, row in shortage_data.iterrows():
                shortage_items.append({
                    'Part Number': row['PartNumber'],
                    'Qty Required': row['QtyAllocated']
                })
    
    # Combine not available parts and shortage items (avoid duplicates)
    all_na_items = []
    
    # First, collect all shortage part numbers to avoid duplicates
    shortage_part_numbers = set()
    for item in shortage_items:
        shortage_part_numbers.add(item['Part Number'])
    
    # Add shortage items first (these have priority)
    for item in shortage_items:
        all_na_items.append({
            'Part Number': item['Part Number'],
            'Qty Required': item['Qty Required'],
            'Type': 'Shortage'
        })
    
    # Add not available parts only if they're not already in shortage items
    if not not_available_df.empty:
        for _, row in not_available_df.iterrows():
            part_number = row['Part Number']
            if part_number not in shortage_part_numbers:
                all_na_items.append({
                    'Part Number': part_number,
                    'Qty Required': row['Qty Required'],
                    'Type': 'Not Available'
                })
    
    # Display consolidated N/A section if there are any items
    if all_na_items:
        # Track the row number for N/A header
        na_header_row = len(combined_data)
        
        # Add N/A header
        combined_data.append({
            'Part Number': 'N/A',
            'Qty Ordered': '',
            'Unit Price': '',
            'Total Cost': ''
        })
        
        # Add all N/A items and calculate total
        for item in all_na_items:
            qty = item['Qty Required']
            combined_data.append({
                'Part Number': item['Part Number'],
                'Qty Ordered': qty,
                'Unit Price': 'N/A',
                'Total Cost': 'N/A'
            })
            not_available_total_qty += qty
        
        # Track the row number for SUBTOTAL - N/A
        subtotal_na_row = len(combined_data)
        
        # Add N/A subtotal
        combined_data.append({
            'Part Number': 'SUBTOTAL - N/A',
            'Qty Ordered': not_available_total_qty,
            'Unit Price': '',
            'Total Cost': 'N/A'
        })
    
    # Calculate total quantity correctly - should be the total from original order
    # Get total quantity required from the original order
    total_qty_required = order_df['QtyRequired'].sum()
    
    combined_data.append({
        'Part Number': '',
        'Qty Ordered': '',
        'Unit Price': '',
        'Total Cost': ''
    })
    
    # Track the row number for TOTAL QTY
    total_qty_row = len(combined_data)
    
    combined_data.append({
        'Part Number': 'TOTAL QTY',
        'Qty Ordered': total_qty_required,
        'Unit Price': '',
        'Total Cost': ''
    })
    
    # Create DataFrame and write to Excel
    if combined_data:
        combined_df = pd.DataFrame(combined_data)
        combined_df.to_excel(writer, sheet_name='Suppliers', index=False, header=False)
        
        # Clear formatting on specific cells and apply bold formatting to table bottom lines
        workbook = writer.book
        worksheet = workbook['Suppliers']
        
        # Create formatting styles
        clear_fill = PatternFill(fill_type=None)
        bold_font = Font(bold=True)
        border_style = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        
        # Apply bold formatting to all subtotal rows
        for subtotal_row in subtotal_rows:
            for col in range(1, 5):  # Columns A-D
                cell = worksheet.cell(row=subtotal_row + 1, column=col)  # +1 for 1-based indexing only
                cell.fill = clear_fill
                cell.font = bold_font
                cell.border = border_style
        
        # Apply bold formatting to supplier header rows
        for supplier_header_row in supplier_header_rows:
            for col in range(1, 5):  # Columns A-D
                cell = worksheet.cell(row=supplier_header_row + 1, column=col)  # +1 for 1-based indexing only
                cell.fill = clear_fill
                cell.font = bold_font
                cell.border = border_style
        
        # Apply bold formatting to column header rows
        for column_header_row in column_header_rows:
            for col in range(1, 5):  # Columns A-D
                cell = worksheet.cell(row=column_header_row + 1, column=col)  # +1 for 1-based indexing only
                cell.fill = clear_fill
                cell.font = bold_font
                cell.border = border_style
        
        # Apply bold formatting to grand total row
        if grand_total_row:
            for col in range(1, 5):  # Columns A-D
                cell = worksheet.cell(row=grand_total_row + 1, column=col)  # +1 for 1-based indexing only
                cell.fill = clear_fill
                cell.font = bold_font
                cell.border = border_style
        
        # Apply bold formatting to N/A header row
        if na_header_row:
            for col in range(1, 5):  # Columns A-D
                cell = worksheet.cell(row=na_header_row + 1, column=col)  # +1 for 1-based indexing only
                cell.fill = clear_fill
                cell.font = bold_font
                cell.border = border_style
        
        # Apply bold formatting to SUBTOTAL - N/A row
        if subtotal_na_row:
            for col in range(1, 5):  # Columns A-D
                cell = worksheet.cell(row=subtotal_na_row + 1, column=col)  # +1 for 1-based indexing only
                cell.fill = clear_fill
                cell.font = bold_font
                cell.border = border_style
        
        # Apply bold formatting to TOTAL QTY row
        if total_qty_row:
            for col in range(1, 5):  # Columns A-D
                cell = worksheet.cell(row=total_qty_row + 1, column=col)  # +1 for 1-based indexing only
                cell.fill = clear_fill
                cell.font = bold_font
                cell.border = border_style
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Add padding and cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
                
    else:
        # Create empty sheet if no data
        empty_df = pd.DataFrame({'Message': ['No supplier data available']})
        empty_df.to_excel(writer, sheet_name='Suppliers', index=False)
        
        # Auto-adjust column widths for empty sheet
        workbook = writer.book
        worksheet = workbook['Suppliers']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Add padding and cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width

# -----------------------------
# File Uploads Section
# -----------------------------
st.header("📥 File Upload")
col1, col2 = st.columns(2)

with col1:
    order_file = st.file_uploader("Upload Order List (Excel)", type=["xlsx"], key="order_upload")
    if order_file:
        st.success(f"✅ Order file uploaded: {order_file.name}")

with col2:
    quote_files = st.file_uploader(
        "Upload Supplier Quotes (Excel) — Multiple files supported", 
        type=["xlsx"], 
        accept_multiple_files=True,
        key="quote_upload"
    )
    if quote_files:
        st.success(f"✅ {len(quote_files)} supplier quote file(s) uploaded")
        for file in quote_files:
            st.write(f"  • {file.name}")

if order_file and quote_files:
    try:
        # -----------------------------
        # Data Loading and Validation
        # -----------------------------
        with st.spinner("Loading and validating data..."):
            orders = pd.read_excel(order_file)
            
            # Combine all quote files
            quotes_list = []
            for quote_file in quote_files:
                df = pd.read_excel(quote_file)
                # Always use filename as supplier name (override any existing Supplier column)
                df['Supplier'] = quote_file.name.split('.')[0]
                quotes_list.append(df)
            
            quotes_df = pd.concat(quotes_list, ignore_index=True)

            # Normalize columns - handle common variations
            def normalize_column_names(df):
                """Normalize column names to handle common variations"""
                column_mapping = {}
                for col in df.columns:
                    clean_col = col.strip()
                    # Handle common variations
                    if clean_col.lower() in ['part_number', 'part number', 'partnumber', 'part']:
                        column_mapping[col] = 'PartNumber'
                    elif clean_col.lower() in ['qty_required', 'qty required', 'qtyrequired', 'quantity', 'qty']:
                        column_mapping[col] = 'QtyRequired'
                    elif clean_col.lower() in ['unit_price', 'unit price', 'unitprice', 'price']:
                        column_mapping[col] = 'UnitPrice'
                    elif clean_col.lower() in ['available_qty', 'available qty', 'availableqty', 'available', 'stock']:
                        column_mapping[col] = 'AvailableQty'
                    elif clean_col.lower() in ['supplier', 'vendor', 'company']:
                        column_mapping[col] = 'Supplier'
                    else:
                        column_mapping[col] = clean_col
                return df.rename(columns=column_mapping)
            
            orders = normalize_column_names(orders)
            quotes_df = normalize_column_names(quotes_df)


            for supplier in quotes_df['Supplier'].unique():
                supplier_data = quotes_df[quotes_df['Supplier'] == supplier]
                st.write(f"Supplier: '{supplier}' (type: {type(supplier)}) - {len(supplier_data)} quotes")
                st.write(f"  Sample data: {supplier_data[['PartNumber', 'UnitPrice', 'AvailableQty']].head(2).to_dict('records')}")

            # Validate required columns
            required_order_cols = {"PartNumber", "QtyRequired"}
            required_quote_cols = {"Supplier", "PartNumber", "UnitPrice", "AvailableQty"}

            if not required_order_cols.issubset(orders.columns):
                st.error(f"❌ Order file missing required columns: {required_order_cols - set(orders.columns)}")
                st.stop()
            
            if not required_quote_cols.issubset(quotes_df.columns):
                st.error(f"❌ Quote files missing required columns: {required_quote_cols - set(quotes_df.columns)}")
                st.stop()

        # Ensure data types
        orders['QtyRequired'] = pd.to_numeric(orders['QtyRequired'], errors='coerce').fillna(0)
        quotes_df['UnitPrice'] = pd.to_numeric(quotes_df['UnitPrice'], errors='coerce').fillna(0)
        quotes_df['AvailableQty'] = pd.to_numeric(quotes_df['AvailableQty'], errors='coerce').fillna(0)
        # Fix Supplier column to ensure consistent string format
        quotes_df['Supplier'] = quotes_df['Supplier'].astype(str)

        # Filter out invalid data - including NaN part numbers
        orders = orders[orders['QtyRequired'] > 0]
        orders = orders[orders['PartNumber'].notna()]  # Remove rows with NaN part numbers
        quotes_df = quotes_df[(quotes_df['UnitPrice'] > 0) & (quotes_df['AvailableQty'] > 0)]

        if orders.empty:
            st.error("No valid orders found after filtering.")
            st.stop()
        
        if quotes_df.empty:
            st.error("No valid quotes found after filtering.")
            st.stop()
    
        # -----------------------------
        # Data Processing and Cleaning
        # -----------------------------
        with st.spinner("Processing and cleaning data..."):
            # Ensure data types
            orders['QtyRequired'] = pd.to_numeric(orders['QtyRequired'], errors='coerce').fillna(0)
            quotes_df['UnitPrice'] = pd.to_numeric(quotes_df['UnitPrice'], errors='coerce').fillna(0)
            quotes_df['AvailableQty'] = pd.to_numeric(quotes_df['AvailableQty'], errors='coerce').fillna(0)
            # Fix Supplier column to ensure consistent string format
            quotes_df['Supplier'] = quotes_df['Supplier'].astype(str)

            # Filter out invalid data - including NaN part numbers
            orders = orders[orders['QtyRequired'] > 0]
            orders = orders[orders['PartNumber'].notna()]  # Remove rows with NaN part numbers
            quotes_df = quotes_df[(quotes_df['UnitPrice'] > 0) & (quotes_df['AvailableQty'] > 0)]

            if orders.empty:
                st.error("❌ No valid orders found after filtering.")
                st.stop()
            
            if quotes_df.empty:
                st.error("❌ No valid quotes found after filtering.")
                st.stop()
            
            st.success(f"✅ Data processed: {len(orders)} orders, {len(quotes_df)} quotes from {quotes_df['Supplier'].nunique()} suppliers")

        # -----------------------------
        # Automated Optimization Engine
        # -----------------------------
        st.header("🤖 Automated Optimization")
        
        if st.button("🚀 Run Optimization", type="primary") or st.session_state.processing_complete:
            with st.spinner("Running automated optimization..."):
                allocation = []
                optimization_stats = {
                    'total_parts': len(orders),
                    'fully_allocated': 0,
                    'partially_allocated': 0,
                    'not_available': 0,
                    'total_cost': 0
                }

                for _, order in orders.iterrows():
                    part = order['PartNumber']
                    qty_needed = order['QtyRequired']

                    part_quotes = quotes_df[quotes_df['PartNumber'] == part].copy()

                    if part_quotes.empty:
                        allocation.append({
                            "PartNumber": part,
                            "Supplier": "NOT AVAILABLE",
                            "AllocatedQty": 0,
                            "UnitPrice": 0,
                            "Total": 0,
                            "QtyRequired": qty_needed,
                            "Status": "Not Available"
                        })
                        optimization_stats['not_available'] += 1
                        continue

                    # Sort by price for optimal allocation
                    part_quotes.sort_values(by='UnitPrice', inplace=True)
                    original_qty_needed = qty_needed

                    for _, q in part_quotes.iterrows():
                        if qty_needed <= 0:
                            break
                        alloc_qty = min(qty_needed, q['AvailableQty'])
                        if alloc_qty > 0:
                            total_cost = alloc_qty * q['UnitPrice']
                            allocation.append({
                                "PartNumber": part,
                                "Supplier": q['Supplier'],
                                "AllocatedQty": alloc_qty,
                                "UnitPrice": q['UnitPrice'],
                                "Total": total_cost,
                                "QtyRequired": original_qty_needed,
                                "Status": "Allocated"
                            })
                            optimization_stats['total_cost'] += total_cost
                            qty_needed -= alloc_qty

                    # Track allocation status
                    if qty_needed == 0:
                        optimization_stats['fully_allocated'] += 1
                    elif qty_needed < original_qty_needed:
                        optimization_stats['partially_allocated'] += 1
                        # Add shortage record
                        allocation.append({
                            "PartNumber": part,
                            "Supplier": "SHORTAGE",
                            "AllocatedQty": qty_needed,
                            "UnitPrice": 0,
                            "Total": 0,
                            "QtyRequired": original_qty_needed,
                            "Status": "Shortage"
                        })
                    else:
                        optimization_stats['not_available'] += 1

                st.session_state.optimized_allocation = allocation
                st.session_state.processing_complete = True
                
                # Display optimization results
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Parts", optimization_stats['total_parts'])
                with col2:
                    st.metric("Fully Allocated", optimization_stats['fully_allocated'])
                with col3:
                    st.metric("Partially Allocated", optimization_stats['partially_allocated'])
                with col4:
                    st.metric("Not Available", optimization_stats['not_available'])
                
                st.metric("💰 Total Optimized Cost", f"${optimization_stats['total_cost']:,.2f}")

        if st.session_state.processing_complete:

            # -----------------------------
            # Excel-Style Interactive Interface
            # -----------------------------
            st.header("📊 Manual Override & Review")
            st.info("💡 **Instructions:** Click on any supplier price to manually select that supplier for the part. The system will automatically handle remaining quantities based on price optimization.")
            
            # Get all unique suppliers from quotes
            all_suppliers = sorted(quotes_df['Supplier'].unique())
            
            # Function to create Excel-style DataFrame for display
            def create_excel_dataframe():
                excel_rows = []
                
                # Process each part to create Excel-style rows
                for part in orders['PartNumber'].unique():
                    # Get the quantity required for this part
                    part_orders = orders[orders['PartNumber'] == part]['QtyRequired']
                    if part_orders.empty:
                        continue
                    total_qty = part_orders.iloc[0]
                    
                    # Get all quotes for this part
                    part_quotes = quotes_df[quotes_df['PartNumber'] == part].copy()
                    
                    # Create row data starting with part info
                    row_data = {
                        'Part Number': part,
                        'Qty Required': total_qty,
                        'Current Selection': st.session_state.supplier_selections.get(part, 'Auto-Optimized')
                    }
                    
                    # Add each supplier's price and availability as columns
                    for supplier in all_suppliers:
                        supplier_quote = part_quotes[part_quotes['Supplier'] == supplier]
                        if not supplier_quote.empty and supplier_quote.iloc[0]['AvailableQty'] > 0:
                            price = supplier_quote.iloc[0]['UnitPrice']
                            available_qty = supplier_quote.iloc[0]['AvailableQty']
                            row_data[f"{supplier}"] = f"${price:.2f} (Qty: {available_qty})"
                        else:
                            row_data[f"{supplier}"] = "N/A"
                    
                    excel_rows.append(row_data)
                
                return pd.DataFrame(excel_rows) if excel_rows else pd.DataFrame()
            
            # Create the DataFrame (this will be recreated on each rerun)
            excel_df = create_excel_dataframe()
            
            if not excel_df.empty:
                
                # Function to highlight selected supplier prices
                def highlight_selected_prices(df):
                    """Apply highlighting to selected supplier prices"""
                    # Create a DataFrame of the same shape filled with empty strings
                    styles = pd.DataFrame('', index=df.index, columns=df.columns)
                    
                    # Get all supplier columns (numeric columns that aren't 'Qty Required')
                    supplier_columns = [col for col in df.columns if col not in ['Part Number', 'Qty Required', 'Current Selection']]
                    
                    # Highlight selected prices
                    for idx, row in df.iterrows():
                        part_number = row['Part Number']
                        current_selection = st.session_state.supplier_selections.get(part_number, 'Auto-Optimized')
                        
                        # If a specific supplier is selected, highlight that column
                        if current_selection != 'Auto-Optimized' and current_selection in supplier_columns:
                            # Highlight the selected supplier's price with light green background
                            styles.loc[idx, current_selection] = 'background-color: #90EE90; font-weight: bold; border: 2px solid #32CD32;'
                    
                    return styles
                
                # Display the interactive table
                st.subheader("🔧 Interactive Supplier Selection")
                
                # Create tabs for different views
                tab1, tab2, tab3 = st.tabs(["📋 Selection Table", "📈 Cost Analysis", "📦 Allocation Summary"])
                
                with tab1:
                    # Apply styling to highlight selected prices
                    styled_df = excel_df.style.apply(highlight_selected_prices, axis=None)
                    
                    # Display the interactive table with styling
                    event = st.dataframe(
                        styled_df,
                        width='stretch',
                        on_select="rerun",
                        selection_mode="single-cell",
                        key="supplier_selection_table"
                    )
                    
                    # Handle cell selection for manual override
                    if event is not None and hasattr(event, 'selection') and event.selection is not None:
                        selection_data = event.selection
                        
                        # Handle the actual selection data structure from Streamlit
                        if hasattr(selection_data, 'cells') and selection_data.cells:
                            try:
                                # Get the first selected cell
                                cell_data = selection_data.cells[0]
                                row_idx = cell_data[0]
                                col_identifier = cell_data[1]
                                
                                # Handle both column index (int) and column name (str) cases
                                if isinstance(col_identifier, str):
                                    # Streamlit returned column name directly
                                    col_name = col_identifier
                                    if col_name in excel_df.columns:
                                        col_idx = excel_df.columns.get_loc(col_name)
                                    else:
                                        st.stop()
                                else:
                                    # Streamlit returned column index
                                    col_idx = int(col_identifier)
                                    if 0 <= col_idx < len(excel_df.columns):
                                        col_name = excel_df.columns[col_idx]
                                    else:
                                        st.stop()
                                
                                # Ensure row index is within bounds
                                if 0 <= row_idx < len(excel_df):
                                    part_number = excel_df.iloc[row_idx]['Part Number']
                                    # Check if user clicked on a supplier column
                                    # A supplier column is any column that's not one of the fixed columns
                                    fixed_columns = ['Part Number', 'Qty Required', 'Current Selection']
                                    is_supplier_column = col_name not in fixed_columns
                                    if is_supplier_column:
                                        cell_value = excel_df.iloc[row_idx][col_name]
                                        # Only allow selection if supplier has valid quote
                                        if cell_value != "N/A":
                                            # The column name should be the supplier name
                                            supplier_name = str(col_name)
                                            # Update the selection in session state
                                            st.session_state.supplier_selections[part_number] = supplier_name
                                            st.success(f"✅ Selected {supplier_name} for part {part_number}")
                                            
                                            # Force immediate rerun to update the display
                                            st.rerun()
                                        else:
                                            st.warning(f"⚠️ {col_name} has no available quote for part {part_number}")
                                    else:
                                        pass  # Column is not a valid supplier column
                                else:
                                    pass  # Index out of bounds
                            except (IndexError, KeyError, TypeError) as e:
                                pass  # Selection error
                        
                        # Also try the old method in case structure varies
                        elif hasattr(selection_data, 'rows') and hasattr(selection_data, 'columns'):
                            rows = selection_data.rows
                            cols = selection_data.columns
                            if rows and cols and len(rows) > 0 and len(cols) > 0:
                                try:
                                    row_idx = rows[0]
                                    col_idx = cols[0]
                                    
                                    # Ensure indices are within bounds
                                    if 0 <= row_idx < len(excel_df) and 0 <= col_idx < len(excel_df.columns):
                                        col_name = excel_df.columns[col_idx]
                                        part_number = excel_df.iloc[row_idx]['Part Number']
                                        
                                        # Check if user clicked on a supplier column
                                        if col_name not in ['Part Number', 'Qty Required', 'Current Selection'] and col_name in all_suppliers:
                                            cell_value = excel_df.iloc[row_idx][col_name]
                                            
                                            # Only allow selection if supplier has valid quote
                                            if cell_value != "N/A":
                                                # Update the selection in session state
                                                st.session_state.supplier_selections[part_number] = col_name
                                                st.success(f"✅ Selected {col_name} for part {part_number}")
                                                
                                                # Force immediate rerun to update the display
                                                st.rerun()
                                
                                except (IndexError, KeyError, TypeError) as e:
                                    pass  # Selection error
                
                # Show current manual selections summary
                if st.session_state.supplier_selections:
                    st.subheader("📝 Current Manual Selections")
                    selection_summary = []
                    for part, supplier in st.session_state.supplier_selections.items():
                        selection_summary.append({
                            "Part Number": part,
                            "Selected Supplier": supplier
                        })
                    
                    st.dataframe(pd.DataFrame(selection_summary), width='stretch')
                    
                    # Reset all selections button
                    if st.button("🔄 Reset All Selections"):
                        st.session_state.supplier_selections = {}
                        st.success("✅ All manual selections have been reset")
                        st.rerun()
                
                with tab2:
                    # Cost analysis comparison
                    st.subheader("💰 Cost Analysis")
                    
                    # Calculate costs for different scenarios
                    auto_cost = sum([item['Total'] for item in st.session_state.optimized_allocation if item['Status'] == 'Allocated'])
                    
                    # Calculate manual override cost
                    manual_cost = 0
                    manual_allocation = []
                    
                    for part in orders['PartNumber'].unique():
                        part_orders = orders[orders['PartNumber'] == part]['QtyRequired']
                        if part_orders.empty:
                            continue
                        qty_needed = part_orders.iloc[0]
                        
                        if part in st.session_state.supplier_selections:
                            selected_supplier = st.session_state.supplier_selections[part]
                            part_quotes = quotes_df[(quotes_df['PartNumber'] == part) & (quotes_df['Supplier'] == selected_supplier)]
                            
                            if not part_quotes.empty:
                                supplier_quote = part_quotes.iloc[0]
                                available_qty = supplier_quote['AvailableQty']
                                unit_price = supplier_quote['UnitPrice']
                                
                                if available_qty >= qty_needed:
                                    # Full allocation to selected supplier
                                    manual_cost += qty_needed * unit_price
                                    manual_allocation.append({
                                        'Part': part,
                                        'Supplier': selected_supplier,
                                        'Qty': qty_needed,
                                        'Cost': qty_needed * unit_price,
                                        'Type': 'Manual Selection'
                                    })
                                else:
                                    # Partial allocation to selected supplier, rest optimized
                                    manual_cost += available_qty * unit_price
                                    manual_allocation.append({
                                        'Part': part,
                                        'Supplier': selected_supplier,
                                        'Qty': available_qty,
                                        'Cost': available_qty * unit_price,
                                        'Type': 'Manual Selection (Partial)'
                                    })
                                    
                                    # Find cheapest option for remaining quantity
                                    remaining_qty = qty_needed - available_qty
                                    remaining_quotes = quotes_df[
                                        (quotes_df['PartNumber'] == part) & 
                                        (quotes_df['Supplier'] != selected_supplier) &
                                        (quotes_df['AvailableQty'] > 0)
                                    ].sort_values('UnitPrice')
                                    
                                    for _, quote in remaining_quotes.iterrows():
                                        if remaining_qty <= 0:
                                            break
                                        alloc_qty = min(remaining_qty, quote['AvailableQty'])
                                        manual_cost += alloc_qty * quote['UnitPrice']
                                        manual_allocation.append({
                                            'Part': part,
                                            'Supplier': quote['Supplier'],
                                            'Qty': alloc_qty,
                                            'Cost': alloc_qty * quote['UnitPrice'],
                                            'Type': 'Auto-Optimized (Remaining)'
                                        })
                                        remaining_qty -= alloc_qty
                    
                    # Display cost comparison
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("🤖 Auto-Optimized Cost", f"${auto_cost:,.2f}")
                    with col2:
                        st.metric("✋ Manual Override Cost", f"${manual_cost:,.2f}")
                    with col3:
                        cost_diff = manual_cost - auto_cost
                        st.metric("📊 Cost Difference", f"${cost_diff:,.2f}", delta=f"{cost_diff:,.2f}")
                
                with tab3:
                    # Final allocation summary
                    st.subheader("📦 Final Allocation Summary")
                    
                    if manual_allocation:
                        manual_df = pd.DataFrame(manual_allocation)
                        st.dataframe(manual_df, width='stretch')
                        
                        # Group by supplier for download preparation
                        supplier_summary = manual_df.groupby('Supplier').agg({
                            'Qty': 'sum',
                            'Cost': 'sum'
                        }).reset_index()
                        
                        st.subheader("📋 Summary by Supplier")
                        st.dataframe(supplier_summary, width='stretch')
        # -----------------------------
        # Download Functionality
        # -----------------------------
        st.header("📥 Download Optimized Orders")
        
        if st.button("📊 Generate Final Allocation", type="primary"):
            # Generate final allocation based on manual selections and auto-optimization
            final_allocation = []
            
            # DEBUG: Show current manual selections
            for part in orders['PartNumber'].unique():
                part_orders = orders[orders['PartNumber'] == part]['QtyRequired']
                if part_orders.empty:
                    continue
                qty_needed = part_orders.iloc[0]
                original_qty_needed = qty_needed
                
                if part in st.session_state.supplier_selections:
                    # Manual selection exists
                    selected_supplier = str(st.session_state.supplier_selections[part])
                    # Convert both supplier values to strings for proper comparison
                    part_quotes = quotes_df[(quotes_df['PartNumber'] == part) & (quotes_df['Supplier'].astype(str) == selected_supplier)]
                    
                    if not part_quotes.empty:
                        supplier_quote = part_quotes.iloc[0]
                        available_qty = supplier_quote['AvailableQty']
                        unit_price = supplier_quote['UnitPrice']
                        
                        if available_qty >= qty_needed:
                            # Full allocation to selected supplier
                            final_allocation.append({
                                'PartNumber': part,
                                'Supplier': selected_supplier,
                                'QtyAllocated': qty_needed,
                                'UnitPrice': unit_price,
                                'TotalCost': qty_needed * unit_price,
                                'AllocationSource': 'Manual Selection'
                            })
                            qty_needed = 0
                        else:
                            # Partial allocation to selected supplier
                            final_allocation.append({
                                'PartNumber': part,
                                'Supplier': selected_supplier,
                                'QtyAllocated': available_qty,
                                'UnitPrice': unit_price,
                                'TotalCost': available_qty * unit_price,
                                'AllocationSource': 'Manual Selection (Partial)'
                            })
                            
                            # Auto-optimize remaining quantity
                            remaining_qty = qty_needed - available_qty
                            remaining_quotes = quotes_df[
                                (quotes_df['PartNumber'] == part) & 
                                (quotes_df['Supplier'].astype(str) != selected_supplier) &
                                (quotes_df['AvailableQty'] > 0)
                            ].sort_values('UnitPrice')
                            
                            for _, quote in remaining_quotes.iterrows():
                                if remaining_qty <= 0:
                                    break
                                alloc_qty = min(remaining_qty, quote['AvailableQty'])
                                final_allocation.append({
                                    'PartNumber': part,
                                    'Supplier': quote['Supplier'],
                                    'QtyAllocated': alloc_qty,
                                    'UnitPrice': quote['UnitPrice'],
                                    'TotalCost': alloc_qty * quote['UnitPrice'],
                                    'AllocationSource': 'Auto-Optimized (Remaining)'
                                })
                                remaining_qty -= alloc_qty
                            qty_needed = remaining_qty
                else:
                    # Use auto-optimization for this part
                    part_quotes = quotes_df[quotes_df['PartNumber'] == part].sort_values('UnitPrice')
                    
                    for _, quote in part_quotes.iterrows():
                        if qty_needed <= 0:
                            break
                        alloc_qty = min(qty_needed, quote['AvailableQty'])
                        if alloc_qty > 0:
                            final_allocation.append({
                                'PartNumber': part,
                                'Supplier': quote['Supplier'],
                                'QtyAllocated': alloc_qty,
                                'UnitPrice': quote['UnitPrice'],
                                'TotalCost': alloc_qty * quote['UnitPrice'],
                                'AllocationSource': 'Auto-Optimized'
                            })
                            qty_needed -= alloc_qty
                
                # Add shortage record if there's unallocated quantity
                if qty_needed > 0:
                    final_allocation.append({
                        'PartNumber': part,
                        'Supplier': 'N/A (SHORTAGE)',
                        'QtyAllocated': qty_needed,
                        'UnitPrice': 0,
                        'TotalCost': 0,
                        'AllocationSource': 'Shortage'
                    })
            
            # Store final allocation in session state
            st.session_state.final_allocation = final_allocation
            
            # Display final allocation summary with color highlighting
            if final_allocation:
                final_df = pd.DataFrame(final_allocation)
                st.subheader("📋 Final Allocation Summary")
                
                # Create a styled dataframe with color highlighting based on AllocationSource
                def highlight_allocation_source(row):
                    """Apply color highlighting based on AllocationSource"""
                    allocation_source = row['AllocationSource']
                    
                    if 'Manual Selection' in allocation_source:
                        if 'Partial' in allocation_source:
                            # Orange for partial manual selections
                            return ['background-color: #FFA500; color: #FF4500; font-weight: bold'] * len(row)
                        else:
                            # Gold for full manual selections
                            return ['background-color: #FFD700; color: #8B4513; font-weight: bold'] * len(row)
                    elif 'Auto-Optimized' in allocation_source:
                        # Light green for auto-optimized
                        return ['background-color: #90EE90; color: #006400; font-weight: bold'] * len(row)
                    else:
                        # Default styling for shortage/other
                        return [''] * len(row)
                
                # Apply styling and display
                styled_df = final_df.style.apply(highlight_allocation_source, axis=1)
                st.dataframe(styled_df, width='stretch')
                
                # Add legend for color coding
                st.markdown("""
                **Color Legend:**
                - 🟡 **Gold**: Manual Selection (Full)
                - 🟠 **Orange**: Manual Selection (Partial)
                - 🟢 **Green**: Auto-Optimized
                """)
                
                # Calculate total cost
                total_cost = final_df['TotalCost'].sum()
                st.metric("💰 Total Final Cost", f"${total_cost:,.2f}")
                
                # Group by supplier for separate order files
                supplier_groups = final_df.groupby('Supplier')
                
                st.subheader("📦 Orders by Supplier")
                
                # Create download buttons for each supplier
                supplier_index = 0  # Add index to ensure unique keys
                for supplier, group in supplier_groups:
                    if supplier not in ['NOT AVAILABLE', 'SHORTAGE', 'N/A (SHORTAGE)']:
                        col1, col2, col3 = st.columns([2, 1, 1])
                        
                        # Convert supplier to string to handle numeric supplier names
                        supplier_str = str(supplier)
                        
                        with col1:
                            st.write(f"**{supplier_str}**")
                            st.write(f"Parts: {len(group)}, Total Cost: ${group['TotalCost'].sum():,.2f}")
                        
                        with col2:
                            # Create CSV for this supplier
                            supplier_csv = group[['PartNumber', 'QtyAllocated', 'UnitPrice', 'TotalCost']].to_csv(index=False)
                            st.download_button(
                                label=f"📄 CSV",
                                data=supplier_csv,
                                file_name=f"order_{supplier_str.lower().replace(' ', '_')}.csv",
                                mime="text/csv",
                                key=f"csv_{supplier_index}_{supplier_str.replace('.', '_').replace(' ', '_')}"
                            )
                        
                        with col3:
                            # Create Excel for this supplier
                            import io
                            excel_buffer = io.BytesIO()
                            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                # Sanitize supplier name for sheet title
                                clean_supplier_name = supplier_str.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_').replace('(', '_').replace(')', '_')
                                sheet_name = clean_supplier_name[:30]
                                
                                group[['PartNumber', 'QtyAllocated', 'UnitPrice', 'TotalCost']].to_excel(
                                    writer, sheet_name=sheet_name, index=False
                                )
                                
                                # Auto-adjust column widths
                                worksheet = writer.sheets[sheet_name]
                                for column in worksheet.columns:
                                    max_length = 0
                                    column_letter = column[0].column_letter
                                    for cell in column:
                                        try:
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = min(max_length + 2, 50)
                                    worksheet.column_dimensions[column_letter].width = adjusted_width
                            excel_buffer.seek(0)
                            
                            st.download_button(
                                label=f"📊 Excel",
                                data=excel_buffer.getvalue(),
                                file_name=f"order_{supplier_str.lower().replace(' ', '_')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"excel_{supplier_index}_{supplier_str.replace('.', '_').replace(' ', '_')}"
                            )
                        
                        supplier_index += 1  # Increment index for next supplier
                
                # Master file download
                st.subheader("📋 Master Files")
                col1, col2 = st.columns(2)
                
                with col1:
                    # Master CSV
                    master_csv = final_df.to_csv(index=False)
                    st.download_button(
                        label="📄 Download Master CSV",
                        data=master_csv,
                        file_name="master_allocation.csv",
                        mime="text/csv"
                    )
                
                with col2:
                    # Enhanced Multi-Sheet Excel
                    master_excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(master_excel_buffer, engine='openpyxl') as writer:
                        
                        # Ensure we have data to work with
                        if final_df.empty:
                            # Create a basic sheet if no allocation data
                            empty_df = pd.DataFrame({'Message': ['No allocation data available. Please run optimization first.']})
                            empty_df.to_excel(writer, sheet_name='No Data', index=False)
                        else:
                            # Sheet 1: Order List with Price Comparison
                            try:
                                sheet1_data = create_price_comparison_sheet(orders, quotes_df, final_df)
                                sheet1_data.to_excel(writer, sheet_name='Order List', index=False)
                                
                                # Apply highlighting to Sheet 1
                                apply_excel_highlighting(writer, 'Order List', sheet1_data, final_df)
                            except Exception as e:
                                st.warning(f"Could not create Order List sheet: {e}")
                                # Create fallback sheet
                                fallback_df = pd.DataFrame({'Error': [f'Order List creation failed: {e}']})
                                fallback_df.to_excel(writer, sheet_name='Order List Error', index=False)
                            
                            # Sheet 2: Combined Suppliers (includes all suppliers and not available items)
                            try:
                                create_combined_suppliers_sheet(writer, final_df, quotes_df, orders)
                            except Exception as e:
                                st.warning(f"Could not create Suppliers sheet: {e}")
                                # Create fallback sheet
                                fallback_df = pd.DataFrame({'Error': [f'Suppliers sheet creation failed: {e}']})
                                fallback_df.to_excel(writer, sheet_name='Suppliers Error', index=False)
                    
                    master_excel_buffer.seek(0)
                    st.download_button(
                        label="📊 Download Enhanced Excel Report",
                        data=master_excel_buffer.getvalue(),
                        file_name="supplier_quote_analysis.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

    except Exception as e:
        st.error(f"❌ Error processing files: {str(e)}")
        st.error("Please check your file formats and try again.")

# -----------------------------
# Footer
# -----------------------------
st.markdown("---")
st.markdown("### 📋 Instructions")
st.markdown("""
1. **Upload Files**: Upload your order list and supplier quote files
2. **Run Optimization**: Click 'Run Optimization' to get automated allocation
3. **Manual Override**: Click on supplier prices in the table to manually select preferred suppliers
4. **Review Costs**: Check the cost analysis to compare auto vs manual selections
5. **Download**: Generate and download optimized order files grouped by supplier
""")

st.markdown("### 📊 File Format Requirements")
st.markdown("""
**Order List CSV/Excel:**
- Columns: `PartNumber`, `QtyRequired`

**Supplier Quote CSV/Excel:**
- Columns: `PartNumber`, `Supplier`, `UnitPrice`, `AvailableQty`
""")